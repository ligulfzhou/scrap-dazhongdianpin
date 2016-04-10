import openpyxl
from openpyxl.styles import Alignment
import pymongo

mongodb = pymongo.MongoClient('127.0.0.1:27017')


def export_xlsx(data, export_filename):
    '''data = [{'sheetname': '总览', titles: ['1', '2', '3'], data: [{1,2,3}, {2,3,4}, {3,4,5}]}]'''
    assert isinstance(data, list)
    wb = openpyxl.Workbook()
    # 这个alignment和上一个的style功能一样
    alignment = Alignment(wrap_text=True)
    ws_num = len(data)
    wss = []
    max_len = []   # [[], [], []], 每一个sheet, 每一列的最大长度, 导出时, 显示更正常
    for i in range(ws_num):
        if i == 0:
            # 第一个sheet是这么取的, 如果直接create_sheet, 生成的第一个sheet是空的
            ws = wb.active
            ws.title = data[i].get('sheetname', '')
        else:
            ws = wb.create_sheet(data[i].get('sheetname', ''))
        for idx, title in enumerate(data[i].get('titles', [])):
            # 这里是ws['A1'] = **, ws[A2] = **
            col = ord('A') + idx
            ws['%s1' % chr(col)] = title     # 写入标题
        # 初始化每一个sheet,每一列最大宽度为0
        max_len.append([0 for _ in range(len(data[i].get('titles', [])))])
        wss.append(ws)
    for idx, ws in enumerate(wss):
        data_lines = data[idx].get('data', [])
        for line, data_line in enumerate(data_lines):
            for col, data_col in enumerate(data_line):
                cur_col = ord('A') + col
                ws['%s%s' % (chr(cur_col), line+1)] = data_col
                ws['%s%s' % (chr(cur_col), line+1)].alignment = alignment
                if len(bytes(data_col, 'GBK')) > max_len[idx][col]:
                    max_len[idx][col] = len(bytes(data_col, 'GBK'))
        for colidx in range(len(data[idx].get('titles', []))):
            cur_col = ord('A') + colidx
            ws.column_dimensions['%s' % chr(cur_col)].width = max_len[idx][colidx]
    wb.save('%s.xlsx' % export_filename)


if __name__ == '__main__':
    xihuqu = mongodb.xihuqu
    stores = xihuqu.store.find({}, {'_id': 0})
    data = [{'sheetname': '商店', 'titles': ['健身馆', '类别', '电话'], 'data': []}]
    for item in stores:
        data[0].get('data').append([item['name'], item['category'], ','.join(item['mobile'])])
    export_xlsx(data, 'xihuqu')

    yuhangqu = mongodb.yuhangqu
    stores = yuhangqu.store.find({}, {'_id': 0})
    data = [{'sheetname': '商店', 'titles': ['健身馆', '类别', '电话'], 'data': []}]
    for item in stores:
        data[0].get('data').append([item['name'], item['category'], ','.join(item['mobile'])])
    export_xlsx(data, 'yuhangqu')

    zhoushan = mongodb.zhoushan
    stores = zhoushan.store.find({}, {'_id': 0})
    data = [{'sheetname': '商店', 'titles': ['健身馆', '类别', '电话'], 'data': []}]
    for item in stores:
        data[0].get('data').append([item['name'], item['category'], ','.join(item['mobile'])])
    export_xlsx(data, 'zhoushan')

    quzhou = mongodb.quzhou
    stores = quzhou.store.find({}, {'_id': 0})
    data = [{'sheetname': '商店', 'titles': ['健身馆', '类别', '电话'], 'data': []}]
    for item in stores:
        data[0].get('data').append([item['name'], item['category'], ','.join(item['mobile'])])
    export_xlsx(data, 'quzhou')
